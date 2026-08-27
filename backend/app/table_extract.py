"""表格照片 → xlsx。视觉模型输出带合并信息的 JSON，openpyxl 还原成有线框、有合并单元格、
可选 A4/A3 纸张的 Excel。跟 OCR 一样是"读取/转录"，不接敏感文件检测。"""
import io
import json
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

TABLE_INSTRUCTION = (
    "识别图片里的表格，尽量还原它的结构（含合并单元格），输出为 JSON 对象：\n"
    '{"rows": 总行数, "cols": 总列数, "cells": ['
    '{"r": 行号, "c": 列号, "rs": 跨行数, "cs": 跨列数, "t": "单元格文字"}, ...]}\n'
    "- r/c 从 0 开始。合并的单元格只输出左上角那一个，rs/cs 填跨度；不合并的 rs=1、cs=1。\n"
    "- 每个非空单元格都要输出；空单元格可以不输出。\n"
    "- 表格上方/外面的大标题（例如某某验收单、某某明细表）也要，作为第 0 行、跨满所有列的合并单元格。\n"
    "- 表格下方的落款/签字行（负责人、日期之类）也一并纳入对应的行。\n"
    "- 只输出能被 JSON.parse 的对象本身，不要 markdown、不要 ```、不要任何解释。\n"
    "图里没有表格就输出 {\"rows\":0,\"cols\":0,\"cells\":[]}。"
)

_PAPER = {"A4": 9, "A3": 8}


def parse_spec(content: str) -> dict:
    s = re.sub(r"^```(?:json)?\s*|\s*```$", "", content.strip(), flags=re.IGNORECASE).strip()
    for cand in (s, s[s.find("{"): s.rfind("}") + 1] if "{" in s and "}" in s else ""):
        if not cand:
            continue
        try:
            data = json.loads(cand)
        except json.JSONDecodeError:
            continue
        if isinstance(data, dict) and "cells" in data:
            return data
        if isinstance(data, list):  # 兜底：模型给了纯二维数组
            cells = [
                {"r": r, "c": c, "rs": 1, "cs": 1, "t": "" if v is None else str(v)}
                for r, row in enumerate(data)
                for c, v in enumerate(row if isinstance(row, list) else [row])
            ]
            rows = len(data)
            cols = max((len(row) for row in data if isinstance(row, list)), default=1)
            return {"rows": rows, "cols": cols, "cells": cells}
    raise ValueError("未能从图片里解析出表格结构")


def build_xlsx(spec: dict, paper: str = "A4", orientation: str = "auto") -> bytes:
    cells = spec.get("cells") or []
    if not cells:
        raise ValueError("图片里没有识别到表格内容")
    rows = int(spec.get("rows") or max(c.get("r", 0) + c.get("rs", 1) for c in cells))
    cols = int(spec.get("cols") or max(c.get("c", 0) + c.get("cs", 1) for c in cells))
    rows = max(1, min(rows, 400))
    cols = max(1, min(cols, 60))

    wb = Workbook()
    ws = wb.active
    ws.title = "表格"
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 先给整张网格铺线框 + 居中自动换行（合并前铺，合并后 MergedCell 不能再改样式）
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = border
            cell.alignment = align

    col_text_len = [4] * (cols + 1)
    for cd in cells:
        r, c = int(cd.get("r", 0)) + 1, int(cd.get("c", 0)) + 1
        if not (1 <= r <= rows and 1 <= c <= cols):
            continue
        rs, cs = max(1, int(cd.get("rs", 1))), max(1, int(cd.get("cs", 1)))
        text = str(cd.get("t", ""))
        ws.cell(row=r, column=c, value=text)
        if r == 1:
            # 第一行如果是跨满整表的一个格，多半是大标题——加大字号
            big = cs >= max(2, cols - 1)
            ws.cell(row=r, column=c).font = Font(bold=True, size=15 if big else 11)
        if rs > 1 or cs > 1:
            try:
                ws.merge_cells(start_row=r, start_column=c,
                               end_row=min(r + rs - 1, rows), end_column=min(c + cs - 1, cols))
            except ValueError:
                pass
        if cs == 1 and c <= cols:
            col_text_len[c] = max(col_text_len[c], len(text))

    for c in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = min(max(col_text_len[c] * 1.8 + 2, 8), 40)

    landscape = orientation == "landscape" or (orientation == "auto" and cols >= 8)
    ws.page_setup.paperSize = _PAPER.get(paper, 9)
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
